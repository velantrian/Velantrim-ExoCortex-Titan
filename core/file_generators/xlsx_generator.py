"""
📊 XLSX Generator v1.0 — Excel через openpyxl
==============================================
Профессиональные Excel-таблицы из GenerationSpec.

Стратегия рендеринга:
- TableBlock → отдельный лист в workbook
- HeadingBlock → название листа
- FactBlock'и группируются в один лист "Facts"
- ParagraphBlock → лист "Notes"

Особенности:
- Цветовая палитра под тему
- Замороженные header'ы
- Auto-fit ширины колонок
- Conditional formatting для confidence
- Data validation для epistemic_state
- Сводный лист "Summary" с метриками
"""

import logging
import time

from .base import (
    FactBlock,
    FileGenerator,
    GenerationResult,
    GenerationSpec,
    HeadingBlock,
    ParagraphBlock,
    TableBlock,
    get_theme,
)

logger = logging.getLogger("velantrim.generators.xlsx")


def _check_available(module_name: str) -> bool:
    import importlib.util
    return importlib.util.find_spec(module_name) is not None


OPENPYXL_AVAILABLE = _check_available("openpyxl")


class XLSXGenerator(FileGenerator):
    format_name = "xlsx"

    def supported_extensions(self) -> list[str]:
        return [".xlsx"]

    def generate(
        self,
        spec: GenerationSpec,
        output_path: str,
    ) -> GenerationResult:
        result = GenerationResult(output_path=output_path, format="xlsx")
        start = time.time()

        if not OPENPYXL_AVAILABLE:
            result.error = "openpyxl не установлен. pip install openpyxl"
            return result

        self._ensure_output_dir(output_path)
        theme = get_theme(spec.theme)

        try:
            sheet_count = self._render(spec, output_path, theme)
            result.method = "openpyxl"
            result.page_count = sheet_count
            result.block_count = len(spec.blocks)
            result.file_size_bytes = self._file_size(output_path)
            result.metadata = self._build_provenance(output_path)
        except Exception as exc:
            logger.error(f"XLSX generation error: {exc}")
            result.error = str(exc)

        result.generation_time_ms = (time.time() - start) * 1000
        return result

    def _render(self, spec: GenerationSpec, output_path: str, theme) -> int:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Border,
            Font,
            PatternFill,
            Side,
        )

        wb = Workbook()
        wb.remove(wb.active)  # удаляем default sheet

        # Метаданные workbook
        wb.properties.title = spec.metadata.title
        wb.properties.creator = spec.metadata.author
        wb.properties.subject = spec.metadata.subject

        # ─── Стили ───
        header_font = Font(
            name=theme.font_heading,
            size=theme.size_md,
            bold=True,
            color=theme.background,
        )
        header_fill = PatternFill(
            start_color=theme.primary,
            end_color=theme.primary,
            fill_type="solid",
        )
        body_font = Font(name=theme.font_body, size=theme.size_sm, color=theme.text)
        border_thin = Border(
            left=Side(style="thin", color=theme.border),
            right=Side(style="thin", color=theme.border),
            top=Side(style="thin", color=theme.border),
            bottom=Side(style="thin", color=theme.border),
        )

        sheet_count = 0

        # ─── Группировка ───
        # Собираем таблицы, факты и параграфы по типам
        tables: list[TableBlock] = []
        facts: list[FactBlock] = []
        paragraphs: list[ParagraphBlock] = []
        current_heading = ""

        for block in spec.blocks:
            if isinstance(block, HeadingBlock):
                current_heading = block.text
            elif isinstance(block, TableBlock):
                # Каждая таблица — отдельный лист
                ws_name = (block.caption or current_heading or f"Таблица {len(tables)+1}")[:31]
                ws = wb.create_sheet(self._safe_sheet_name(ws_name, wb))
                self._render_table_sheet(
                    ws, block, theme,
                    header_font, header_fill, body_font, border_thin,
                )
                tables.append(block)
                sheet_count += 1
            elif isinstance(block, FactBlock):
                facts.append(block)
            elif isinstance(block, ParagraphBlock):
                paragraphs.append(block)

        # ─── Сводный лист Summary ───
        if facts or tables or paragraphs:
            summary_ws = wb.create_sheet("📊 Summary", 0)  # вставляем первым
            self._render_summary_sheet(
                summary_ws, spec, facts, tables, theme,
                header_font, header_fill, body_font,
            )
            sheet_count += 1

        # ─── Лист фактов ───
        if facts:
            facts_ws = wb.create_sheet("🔱 Facts")
            self._render_facts_sheet(
                facts_ws, facts, theme,
                header_font, header_fill, body_font, border_thin,
            )
            sheet_count += 1

        # ─── Лист заметок ───
        if paragraphs:
            notes_ws = wb.create_sheet("📝 Notes")
            self._render_notes_sheet(
                notes_ws, paragraphs, theme, body_font,
            )
            sheet_count += 1

        wb.save(output_path)
        return sheet_count

    # ─── Sheets ───────────────────────────────────────────────────────────────

    def _render_table_sheet(
        self, ws, block: TableBlock, theme,
        header_font, header_fill, body_font, border_thin,
    ) -> None:
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        # Headers
        for col_idx, header in enumerate(block.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
            cell.border = border_thin

        # Body
        for row_idx, row in enumerate(block.rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.border = border_thin
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Заморозка header row
        ws.freeze_panes = "A2"

        # Auto-fit ширины
        for col_idx, header in enumerate(block.headers, start=1):
            letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row in block.rows:
                if col_idx <= len(row):
                    val_len = len(str(row[col_idx - 1])) if row[col_idx - 1] else 0
                    max_len = max(max_len, val_len)
            ws.column_dimensions[letter].width = min(max_len + 4, 60)

    def _render_facts_sheet(
        self, ws, facts: list[FactBlock], theme,
        header_font, header_fill, body_font, border_thin,
    ) -> None:
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        headers = [
            "Fact ID", "Claim", "Confidence", "Epistemic State", "Source"
        ]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border_thin

        for row_idx, fact in enumerate(facts, start=2):
            ws.cell(row=row_idx, column=1, value=fact.fact_id).font = body_font
            ws.cell(row=row_idx, column=2, value=fact.claim).font = body_font
            ws.cell(row=row_idx, column=3, value=fact.confidence).font = body_font
            ws.cell(row=row_idx, column=4, value=fact.epistemic_state).font = body_font
            ws.cell(row=row_idx, column=5, value=fact.source).font = body_font

            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).border = border_thin

        # Conditional formatting для confidence (column 3)
        # >= 0.9 → зелёный, 0.5-0.9 → жёлтый, < 0.5 → красный
        green_fill = PatternFill(
            start_color=theme.success, end_color=theme.success,
            fill_type="solid",
        )
        yellow_fill = PatternFill(
            start_color=theme.warning, end_color=theme.warning,
            fill_type="solid",
        )
        red_fill = PatternFill(
            start_color=theme.danger, end_color=theme.danger,
            fill_type="solid",
        )

        if len(facts) > 0:
            range_str = f"C2:C{len(facts)+1}"
            ws.conditional_formatting.add(
                range_str,
                CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=green_fill),
            )
            ws.conditional_formatting.add(
                range_str,
                CellIsRule(operator="between", formula=["0.5", "0.899"], fill=yellow_fill),
            )
            ws.conditional_formatting.add(
                range_str,
                CellIsRule(operator="lessThan", formula=["0.5"], fill=red_fill),
            )

        # Freeze + widths
        ws.freeze_panes = "A2"
        widths = [20, 60, 14, 18, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _render_summary_sheet(
        self, ws, spec: GenerationSpec, facts: list[FactBlock],
        tables: list[TableBlock], theme,
        header_font, header_fill, body_font,
    ) -> None:
        from openpyxl.styles import Font

        # Title
        ws["A1"] = spec.metadata.title
        ws["A1"].font = Font(
            name=theme.font_heading,
            size=theme.size_2xl, bold=True,
            color=theme.primary,
        )
        ws.merge_cells("A1:D1")

        # Метаданные
        ws["A3"] = "Автор:"
        ws["B3"] = spec.metadata.author
        ws["A4"] = "Создан:"
        ws["B4"] = spec.metadata.created or "—"
        ws["A5"] = "Описание:"
        ws["B5"] = spec.metadata.description

        # Статистика
        ws["A7"] = "📊 Статистика"
        ws["A7"].font = Font(
            name=theme.font_heading,
            size=theme.size_xl, bold=True,
            color=theme.primary,
        )
        ws.merge_cells("A7:D7")

        stats = [
            ("Всего блоков", len(spec.blocks)),
            ("Таблиц", len(tables)),
            ("Фактов", len(facts)),
            ("Тема", spec.theme),
        ]
        for i, (k, v) in enumerate(stats, start=9):
            ws.cell(row=i, column=1, value=k).font = body_font
            ws.cell(row=i, column=2, value=str(v)).font = Font(
                name=theme.font_body, size=theme.size_md, bold=True,
            )

        if facts:
            # Распределение по states
            ws.cell(row=9 + len(stats) + 1, column=1, value="📈 Распределение по состояниям")
            ws.cell(row=9 + len(stats) + 1, column=1).font = Font(
                name=theme.font_heading, size=theme.size_lg, bold=True,
                color=theme.primary,
            )

            state_counts: dict[str, int] = {}
            for fact in facts:
                state_counts[fact.epistemic_state] = state_counts.get(fact.epistemic_state, 0) + 1

            for i, (state, count) in enumerate(
                sorted(state_counts.items(), key=lambda x: -x[1]),
                start=9 + len(stats) + 3,
            ):
                ws.cell(row=i, column=1, value=state).font = body_font
                ws.cell(row=i, column=2, value=count).font = body_font

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

    def _render_notes_sheet(
        self, ws, paragraphs: list[ParagraphBlock], theme, body_font,
    ) -> None:
        from openpyxl.styles import Alignment

        ws["A1"] = "📝 Заметки"
        from openpyxl.styles import Font
        ws["A1"].font = Font(
            name=theme.font_heading,
            size=theme.size_xl, bold=True,
            color=theme.primary,
        )

        for i, p in enumerate(paragraphs, start=3):
            cell = ws.cell(row=i, column=1, value=p.text)
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 100

    @staticmethod
    def _safe_sheet_name(name: str, wb) -> str:
        """Excel не разрешает /\\*?[]: в названиях листов и >31 символов."""
        clean = name.replace("/", "_").replace("\\", "_")
        for ch in "*?[]:":
            clean = clean.replace(ch, "_")
        clean = clean[:31]
        # Проверка уникальности
        base = clean
        i = 1
        existing = {s.title for s in wb.worksheets}
        while clean in existing:
            suffix = f" ({i})"
            clean = base[:31 - len(suffix)] + suffix
            i += 1
        return clean
